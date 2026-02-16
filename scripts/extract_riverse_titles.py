"""
RIVERSE_統合コンテンツリスト.xlsx에서 리버스 작품 목록을 추출하여
data/riverse_titles.json 파일로 저장하는 스크립트
"""

import openpyxl
import json
from pathlib import Path


def extract_riverse_titles():
    """XLSX에서 리버스 작품 추출"""
    # 경로 설정
    project_root = Path(__file__).parent.parent
    xlsx_path = project_root / 'docs' / 'RIVERSE_統合コンテンツリスト.xlsx'
    output_path = project_root / 'data' / 'riverse_titles.json'

    print(f"📂 XLSX 파일 읽기: {xlsx_path}")

    # 엑셀 파일 로드
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except FileNotFoundError:
        print(f"❌ 파일을 찾을 수 없습니다: {xlsx_path}")
        return

    riverse_titles = {}

    # 모든 시트 순회
    print(f"📊 시트 개수: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        print(f"\n🔍 시트 처리 중: {sheet_name}")
        sheet = wb[sheet_name]

        # 첫 번째 행에서 컬럼 헤더 찾기
        headers = [cell.value for cell in sheet[1]]

        # 일본어/한국어 제목 컬럼 인덱스 찾기
        jp_col = None
        kr_col = None

        for i, header in enumerate(headers):
            if header:
                header_str = str(header).lower()
                # 일본어 제목 컬럼 찾기 (다양한 표기 지원)
                if any(keyword in header_str for keyword in ['日本', 'タイトル', 'jp', 'japanese']):
                    if jp_col is None:  # 첫 번째 매칭만
                        jp_col = i
                        print(f"  → 일본어 컬럼: {i} ({header})")

                # 한국어 제목 컬럼 찾기
                if any(keyword in header_str for keyword in ['韓国', '한국', '제목', 'kr', 'korean', '작품명']):
                    if kr_col is None:  # 첫 번째 매칭만
                        kr_col = i
                        print(f"  → 한국어 컬럼: {i} ({header})")

        # 데이터 추출
        if jp_col is not None and kr_col is not None:
            count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if row_idx > sheet.max_row:
                    break

                jp_title = row[jp_col].value
                kr_title = row[kr_col].value

                # 제목이 모두 있는 경우만 추가
                if jp_title and kr_title:
                    jp_title = jp_title.strip()
                    kr_title = kr_title.strip()

                    # 빈 문자열이 아닌 경우만
                    if jp_title and kr_title:
                        riverse_titles[jp_title] = kr_title
                        count += 1

            print(f"  ✅ {count}개 작품 추출")
        else:
            print(f"  ⚠️  일본어/한국어 컬럼을 찾을 수 없습니다")

    # 중복 제거 확인
    print(f"\n📝 총 {len(riverse_titles)}개 고유 작품")

    # JSON 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(riverse_titles, f, ensure_ascii=False, indent=2)

    print(f"✅ 저장 완료: {output_path}")
    print(f"\n📊 샘플 데이터 (첫 5개):")
    for i, (jp, kr) in enumerate(list(riverse_titles.items())[:5], 1):
        print(f"  {i}. {jp} → {kr}")

    return riverse_titles


if __name__ == "__main__":
    print("=" * 60)
    print("RIVERSE 작품 리스트 추출 스크립트")
    print("=" * 60)
    extract_riverse_titles()
    print("\n" + "=" * 60)
    print("완료!")
    print("=" * 60)
